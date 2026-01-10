#!/usr/bin/env python3
"""
Import data from Mouse_Mastersheet.xlsx into the database.

Usage:
    python import_data.py [--dry-run]
"""

import argparse
from datetime import datetime
from pathlib import Path

# Use openpyxl directly since pandas has Python 3.14 compatibility issues
from openpyxl import load_workbook

from app import create_app, db
from models import Company, Inventory


def parse_date(value):
    """Parse date from Excel cell."""
    if value is None or value == 'X' or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None


def parse_float(value):
    """Parse float from Excel cell."""
    if value is None or value == '' or value == 'NaN':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_bool(value, true_values=('Y', 'Yes', 'yes', 'y', True)):
    """Parse boolean from Excel cell."""
    return value in true_values


def get_or_create_company(session, name):
    """Get existing company or create new one."""
    if not name or name == 'NaN' or str(name).strip() == '':
        return None

    name = str(name).strip()
    company = session.query(Company).filter_by(name=name).first()
    if not company:
        company = Company(name=name, relationship_status='active')
        session.add(company)
        session.flush()  # Get the ID
    return company


def import_review_units(ws, session, dry_run=False):
    """Import review units from worksheet."""
    print("\n=== Importing Review Units ===")
    imported = 0
    skipped = 0

    # Get headers from row 1
    headers = [cell.value for cell in ws[1]]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(headers, row))

        product_name = row_dict.get('Model')
        if not product_name or str(product_name).strip() == '':
            skipped += 1
            continue

        product_name = str(product_name).strip()

        # Check if already exists
        existing = session.query(Inventory).filter_by(product_name=product_name).first()
        if existing:
            print(f"  Skipping (exists): {product_name}")
            skipped += 1
            continue

        # Get or create company
        source = row_dict.get('Source')
        company = get_or_create_company(session, source)

        # Map category
        category_raw = str(row_dict.get('Category', 'Mouse')).lower()
        category_map = {
            'mouse': 'mouse',
            'keyboard': 'keyboard',
            'iem': 'iem',
            'pad': 'mousepad',
            'mousepad': 'mousepad',
        }
        category = category_map.get(category_raw, 'other')

        # Determine status
        done = parse_bool(row_dict.get('Done?'))
        sold = parse_bool(row_dict.get('Sold?'))
        if sold:
            status = 'sold'
        elif done:
            status = 'reviewed'
        else:
            status = 'in_queue'

        # Create inventory item
        item = Inventory(
            product_name=product_name,
            company_id=company.id if company else None,
            category=category,
            source_type='review_unit',
            date_acquired=parse_date(row_dict.get('Acquired Date')),
            cost=0.0,  # Review units are free
            on_amazon=parse_bool(row_dict.get('Amazon?'), ('Y', 'Yes', 'yes', 'y', 'N' != row_dict.get('Amazon?'))),
            deadline=parse_date(row_dict.get('Deadline')),
            status=status,
            condition='new',
            notes=str(row_dict.get('Notes', '') or '') + (' ' + str(row_dict.get('Notes 2', '') or '')).strip() or None,
            short_url=row_dict.get('Short') if row_dict.get('Short') and row_dict.get('Short') != 'X' else None,
            short_publish_date=parse_date(row_dict.get('Short Date')),
            video_url=row_dict.get('Video') if row_dict.get('Video') and row_dict.get('Video') != 'X' else None,
            video_publish_date=parse_date(row_dict.get('Video Date')),
            sold=sold,
            sale_price=parse_float(row_dict.get('Price')),
            fees=parse_float(row_dict.get('Fees')),
            shipping=parse_float(row_dict.get('Shipping')),
            marketplace=str(row_dict.get('Marketplace', '')).lower() if row_dict.get('Marketplace') else None,
            buyer=row_dict.get('Buyer') if row_dict.get('Buyer') else None,
        )

        if not dry_run:
            session.add(item)

        print(f"  + {product_name} ({category}, {status})")
        imported += 1

    print(f"\nReview Units: {imported} imported, {skipped} skipped")
    return imported


def import_purchased_mice(ws, session, dry_run=False):
    """Import purchased mice from worksheet."""
    print("\n=== Importing Purchased Mice ===")
    imported = 0
    skipped = 0

    headers = [cell.value for cell in ws[1]]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(headers, row))

        product_name = row_dict.get('Model')
        if not product_name or str(product_name).strip() == '':
            skipped += 1
            continue

        product_name = str(product_name).strip()

        # Check if already exists
        existing = session.query(Inventory).filter_by(product_name=product_name).first()
        if existing:
            print(f"  Skipping (exists): {product_name}")
            skipped += 1
            continue

        # Get or create company from Source
        source = row_dict.get('Source')
        company = get_or_create_company(session, source) if source and source not in ('Aliexpress', 'MouseMarket', 'Xraypad') else None

        # Map category
        category_raw = str(row_dict.get('Category', 'Mouse')).lower()
        category_map = {
            'mouse': 'mouse',
            'keyboard': 'keyboard',
            'pad': 'mousepad',
            'mousepad': 'mousepad',
        }
        category = category_map.get(category_raw, 'other')

        # Determine status
        sold = parse_bool(row_dict.get('Sold?'))
        if sold:
            status = 'sold'
        else:
            status = 'keeping'

        cost = parse_float(row_dict.get('Cost')) or 0.0

        item = Inventory(
            product_name=product_name,
            company_id=company.id if company else None,
            category=category,
            source_type='personal_purchase',
            date_acquired=parse_date(row_dict.get('Purchase Date')),
            cost=cost,
            status=status,
            condition='new',
            notes=str(row_dict.get('Notes', '') or '') + (' ' + str(row_dict.get('Notes 2', '') or '')).strip() or None,
            sold=sold,
            sale_price=parse_float(row_dict.get('Price')),
            fees=parse_float(row_dict.get('Fees')),
            shipping=parse_float(row_dict.get('Shipping')),
            marketplace=str(row_dict.get('Marketplace', '')).lower() if row_dict.get('Marketplace') else None,
            buyer=row_dict.get('Buyer') if row_dict.get('Buyer') else None,
        )

        if not dry_run:
            session.add(item)

        print(f"  + {product_name} (${cost:.2f}, {status})")
        imported += 1

    print(f"\nPurchased: {imported} imported, {skipped} skipped")
    return imported


def import_affiliate_companies(ws, session, dry_run=False):
    """Import/update companies from affiliate sheet."""
    print("\n=== Importing Affiliate Companies ===")
    imported = 0
    updated = 0

    headers = [cell.value for cell in ws[1]]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(headers, row))

        company_name = row_dict.get('Company')
        if not company_name or str(company_name).strip() == '':
            continue

        company_name = str(company_name).strip()

        # Check if exists
        company = session.query(Company).filter_by(name=company_name).first()

        affiliate_link = row_dict.get('Affiliate Link')
        code = row_dict.get('Code')

        if company:
            # Update affiliate info
            if affiliate_link and not company.affiliate_link:
                company.affiliate_link = affiliate_link
                company.affiliate_status = 'yes'
                updated += 1
                print(f"  ~ Updated: {company_name}")
        else:
            # Create new company
            company = Company(
                name=company_name,
                category='mice',
                relationship_status='affiliate_only',
                affiliate_status='yes' if affiliate_link else 'no',
                affiliate_link=affiliate_link if affiliate_link else None,
                affiliate_code=code if code else None,
            )
            if not dry_run:
                session.add(company)
            imported += 1
            print(f"  + {company_name}")

    print(f"\nAffiliate Companies: {imported} new, {updated} updated")
    return imported


def main():
    parser = argparse.ArgumentParser(description='Import spreadsheet data')
    parser.add_argument('--dry-run', action='store_true', help='Preview without saving')
    parser.add_argument('--file', default='Mouse_Mastersheet.xlsx', help='Spreadsheet file path')
    args = parser.parse_args()

    spreadsheet_path = Path(args.file)
    if not spreadsheet_path.exists():
        print(f"Error: File not found: {spreadsheet_path}")
        return 1

    print(f"Loading: {spreadsheet_path}")
    if args.dry_run:
        print("*** DRY RUN - No changes will be saved ***\n")

    wb = load_workbook(spreadsheet_path, read_only=True, data_only=True)

    app = create_app()
    with app.app_context():
        total_imported = 0

        # Import review units
        if 'Review Units' in wb.sheetnames:
            total_imported += import_review_units(wb['Review Units'], db.session, args.dry_run)

        # Import purchased mice
        if 'Purchased Mice' in wb.sheetnames:
            total_imported += import_purchased_mice(wb['Purchased Mice'], db.session, args.dry_run)

        # Import affiliate companies
        if 'Affiliate Sales' in wb.sheetnames:
            total_imported += import_affiliate_companies(wb['Affiliate Sales'], db.session, args.dry_run)

        if not args.dry_run:
            db.session.commit()
            print(f"\n=== Done! {total_imported} items imported. ===")
        else:
            print(f"\n=== Dry run complete. Would import {total_imported} items. ===")

    return 0


if __name__ == '__main__':
    exit(main())
